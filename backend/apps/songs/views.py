"""
ViewSets per Song, LyricLine e ChordAnnotation.
Un ViewSet raggruppa tutta la logica CRUD (list, create, retrieve, update, destroy)
in una sola classe. Il router DRF genera automaticamente le URL.
"""
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response


class SongPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 1000

from apps.common.permissions import IsOwner
from .filters import SongFilter
from .models import ChordAnnotation, LyricLine, Song
from .serializers import (
    ChordAnnotationSerializer,
    LyricLineSerializer,
    SongListSerializer,
    SongSerializer,
)
from .services import transpose_song_chords


class SongViewSet(viewsets.ModelViewSet):
    """
    CRUD completo per le canzoni.
    Ogni utente vede e gestisce solo le proprie canzoni (filtro per owner).
    In lista usa SongListSerializer (leggero), in dettaglio SongSerializer (completo).
    Supporta filtro per key/mode e ricerca per title/artist.
    """
    permission_classes = [permissions.IsAuthenticated, IsOwner]
    pagination_class = SongPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = SongFilter
    search_fields = ['title', 'artist']
    ordering_fields = ['title', 'song_number', 'created_at', 'bpm']
    ordering = ['song_number']

    def get_queryset(self):
        from django.db.models import Q
        from apps.groups.models import Membership

        user = self.request.user
        group_id = self.request.query_params.get('group')

        if group_id:
            # Lista canti di un gruppo specifico
            is_member = Membership.objects.filter(group_id=group_id, user=user).exists()
            if not is_member:
                return Song.objects.none()
            return Song.objects.filter(group_id=group_id).prefetch_related('lyric_lines__chords')

        # Dettaglio / lista personale:
        # canti personali (owner=user, group=None) OPPURE canti dei gruppi di cui è membro
        user_group_ids = Membership.objects.filter(user=user).values_list('group_id', flat=True)
        return Song.objects.filter(
            Q(owner=user, group=None) | Q(group_id__in=user_group_ids)
        ).prefetch_related('lyric_lines__chords')

    def get_serializer_class(self):
        """Serializer leggero per la lista, completo per il dettaglio."""
        if self.action == 'list':
            return SongListSerializer
        return SongSerializer

    def perform_create(self, serializer):
        user = self.request.user
        if user.is_guest:
            group_id = self.request.data.get('group')
            if group_id:
                if Song.objects.filter(group_id=group_id).count() >= 100:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied('Limite di 100 canti per gruppo raggiunto. Registrati per continuare.')
            else:
                if Song.objects.filter(owner=user, group=None).count() >= 50:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied('Limite di 50 canti personali raggiunto. Registrati per continuare.')
        serializer.save(owner=user)

    @action(detail=False, methods=['get'], url_path='import-template')
    def import_template(self, request):
        """GET /api/v1/songs/import-template/ — scarica template Excel compilabile."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, fills
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.worksheet.dimensions import RowDimension
        from django.http import HttpResponse

        KEYS = ['C', 'C#', 'Db', 'D', 'D#', 'Eb', 'E', 'F', 'F#', 'Gb', 'G', 'G#', 'Ab', 'A', 'A#', 'Bb', 'B']
        MODES = ['major', 'minor']

        wb = Workbook()

        # ── Foglio 1: Istruzioni ─────────────────────────────────────────────
        ws_help = wb.active
        ws_help.title = 'Istruzioni'
        ws_help.column_dimensions['A'].width = 22
        ws_help.column_dimensions['B'].width = 55
        ws_help.column_dimensions['C'].width = 28

        title_font = Font(bold=True, size=14, color='1F4E79')
        head_font  = Font(bold=True, size=10, color='FFFFFF')
        body_font  = Font(size=10)
        req_fill   = PatternFill('solid', fgColor='C00000')
        opt_fill   = PatternFill('solid', fgColor='2E75B6')
        head_fill  = PatternFill('solid', fgColor='1F4E79')
        thin = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF'),
        )

        ws_help.merge_cells('A1:C1')
        t = ws_help.cell(1, 1, '🎵 Music Platform — Istruzioni importazione canti')
        t.font = title_font
        t.alignment = Alignment(horizontal='left', vertical='center')
        ws_help.row_dimensions[1].height = 28

        ws_help.merge_cells('A2:C2')
        ws_help.cell(2, 1, 'Compila il foglio "Canti" e importalo dall\'app. I campi con * sono obbligatori.')
        ws_help.cell(2, 1).font = Font(size=10, italic=True, color='595959')

        ws_help.row_dimensions[3].height = 6  # spazio

        for col, label in enumerate(['Colonna', 'Descrizione', 'Valori accettati'], 1):
            c = ws_help.cell(4, col, label)
            c.font = head_font; c.fill = head_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin
        ws_help.row_dimensions[4].height = 20

        fields = [
            ('title  ★ OBBLIGATORIO', 'Titolo del canto',                              'Testo libero',                True),
            ('artist',                'Artista o autore',                               'Testo libero',                False),
            ('key',                   'Tonalità (usa il menu a tendina)',                ', '.join(KEYS),               False),
            ('mode',                  'Modo (usa il menu a tendina)',                    'major  oppure  minor',        False),
            ('bpm',                   'Tempo in battiti per minuto',                     'Numero intero es: 120',       False),
            ('time_signature',        'Indicazione di tempo',                            'es: 4/4  3/4  6/8  12/8',    False),
            ('notes',                 'Note libere sul canto',                           'Testo libero',                False),
            ('song_number',           'Numero d\'ordine nel repertorio',                 'Numero intero es: 42',        False),
            ('topics',                'Argomenti / categorie (più valori separati da ,)', 'es: adorazione, lode, Maria', False),
        ]

        req_label_fill  = PatternFill('solid', fgColor='FFE7E7')
        opt_label_fill  = PatternFill('solid', fgColor='EBF3FB')

        for r, (col_name, desc, values, required) in enumerate(fields, start=5):
            bg = req_label_fill if required else opt_label_fill
            badge_fill = req_fill if required else opt_fill
            badge_text = '★ OBB.' if required else 'facolt.'

            c1 = ws_help.cell(r, 1, col_name)
            c1.font = Font(bold=required, size=10); c1.fill = bg; c1.border = thin
            c1.alignment = Alignment(vertical='center', wrap_text=True)

            c2 = ws_help.cell(r, 2, desc)
            c2.font = body_font; c2.fill = bg; c2.border = thin
            c2.alignment = Alignment(vertical='center', wrap_text=True)

            c3 = ws_help.cell(r, 3, values)
            c3.font = Font(size=9, color='595959'); c3.fill = bg; c3.border = thin
            c3.alignment = Alignment(vertical='center', wrap_text=True)

            ws_help.row_dimensions[r].height = 30

        ws_help.merge_cells('A15:C15')
        ws_help.cell(15, 1, '→ Vai al foglio "Canti" per inserire i tuoi dati')
        ws_help.cell(15, 1).font = Font(bold=True, color='2E75B6', size=11)
        ws_help.sheet_view.showGridLines = False

        # ── Foglio 2: Canti ──────────────────────────────────────────────────
        ws = wb.create_sheet('Canti')

        columns = [
            # (intestazione, larghezza, obbligatorio)
            ('title',          32, True),
            ('artist',         26, False),
            ('key',            10, False),
            ('mode',           12, False),
            ('bpm',            10, False),
            ('time_signature', 16, False),
            ('notes',          42, False),
            ('song_number',    14, False),
            ('topics',         42, False),
        ]

        req_hdr_fill = PatternFill('solid', fgColor='C00000')
        opt_hdr_fill = PatternFill('solid', fgColor='2E75B6')
        req_data_fill = PatternFill('solid', fgColor='FFF2F2')
        opt_data_fill = PatternFill('solid', fgColor='F5FAFF')

        for col_idx, (name, width, required) in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            label = f'{name}  ★' if required else name
            cell = ws.cell(1, col_idx, label)
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = req_hdr_fill if required else opt_hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin
            ws.column_dimensions[col_letter].width = width

            # Colora le celle dati (righe 2-500)
            data_fill = req_data_fill if required else opt_data_fill
            for row_idx in range(2, 501):
                dc = ws.cell(row_idx, col_idx)
                dc.fill = data_fill
                dc.border = thin
                dc.alignment = Alignment(vertical='center', wrap_text=(col_idx in (7, 9)))

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'  # intestazione sempre visibile

        # Dropdown key (colonna C)
        dv_key = DataValidation(
            type='list',
            formula1='"' + ','.join(KEYS) + '"',
            allow_blank=True,
            showDropDown=False,
            error='Valore non valido. Scegli dal menu.',
            errorTitle='Tonalità non valida',
            prompt='Scegli la tonalità dal menu',
            promptTitle='Tonalità',
        )
        dv_key.sqref = 'C2:C500'
        ws.add_data_validation(dv_key)

        # Dropdown mode (colonna D)
        dv_mode = DataValidation(
            type='list',
            formula1='"major,minor"',
            allow_blank=True,
            showDropDown=False,
            error='Scrivi "major" oppure "minor".',
            errorTitle='Modo non valido',
            prompt='Scegli "major" o "minor"',
            promptTitle='Modo',
        )
        dv_mode.sqref = 'D2:D500'
        ws.add_data_validation(dv_mode)

        # Riga esempio (precompilata)
        example = ['Ave Maria', 'Schubert', 'F', 'major', 120, '4/4', 'Canto mariano classico', 1, 'adorazione, Maria']
        for col_idx, val in enumerate(example, 1):
            c = ws.cell(2, col_idx, val)
            c.font = Font(italic=True, color='595959', size=9)

        ws.sheet_view.showGridLines = False

        # Imposta "Canti" come foglio attivo all'apertura
        wb.active = ws

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="template_canti.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='import')
    def bulk_import(self, request):
        """POST /api/v1/songs/import/ — importa canti da file Excel."""
        from openpyxl import load_workbook
        from apps.groups.models import MusicGroup, Membership

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Nessun file fornito.'}, status=status.HTTP_400_BAD_REQUEST)

        group_id = request.data.get('group') or None
        group = None
        if group_id:
            try:
                group = MusicGroup.objects.get(pk=group_id)
                if not Membership.objects.filter(group=group, user=request.user).exists():
                    return Response({'detail': 'Non sei membro di questo gruppo.'}, status=status.HTTP_403_FORBIDDEN)
            except MusicGroup.DoesNotExist:
                return Response({'detail': 'Gruppo non trovato.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb['Canti'] if 'Canti' in wb.sheetnames else wb.active
        except Exception:
            return Response({'detail': 'File non valido. Carica un file .xlsx.'}, status=status.HTTP_400_BAD_REQUEST)

        VALID_KEYS = {'C', 'C#', 'Db', 'D', 'D#', 'Eb', 'E', 'F', 'F#', 'Gb', 'G', 'G#', 'Ab', 'A', 'A#', 'Bb', 'B', ''}

        imported = 0
        errors = []
        songs_to_create = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            title = str(row[0]).strip() if row[0] is not None else ''
            if not title or title == 'None':
                errors.append({'row': i, 'error': 'Titolo obbligatorio.'})
                continue

            key = str(row[2]).strip() if row[2] is not None else ''
            if key not in VALID_KEYS:
                errors.append({'row': i, 'error': f'Tonalità non valida: "{key}"'})
                continue

            mode_raw = str(row[3]).strip().lower() if row[3] is not None else ''
            mode = mode_raw if mode_raw in ('major', 'minor') else 'major'

            bpm = None
            if row[4] is not None:
                try:
                    bpm = int(row[4])
                    if bpm <= 0:
                        bpm = None
                except (ValueError, TypeError):
                    pass

            song_number = None
            if row[7] is not None:
                try:
                    song_number = int(row[7])
                except (ValueError, TypeError):
                    pass

            topics_raw = str(row[8]).strip() if row[8] is not None else ''
            topics = [t.strip() for t in topics_raw.split(',') if t.strip()] if topics_raw else []

            songs_to_create.append(Song(
                title=title,
                artist=str(row[1]).strip() if row[1] is not None else '',
                key=key,
                mode=mode,
                bpm=bpm,
                time_signature=str(row[5]).strip() if row[5] is not None else '4/4',
                notes=str(row[6]).strip() if row[6] is not None else '',
                song_number=song_number,
                topics=topics,
                owner=request.user,
                group=group,
            ))

        if songs_to_create:
            Song.objects.bulk_create(songs_to_create)
            imported = len(songs_to_create)

        return Response({'imported': imported, 'errors': errors},
                        status=status.HTTP_201_CREATED if imported > 0 else status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='topics')
    def topics(self, request):
        """GET /api/v1/songs/topics/?group=<id> — lista argomenti disponibili."""
        qs = self.get_queryset()
        all_topics = set()
        for topics_list in qs.values_list('topics', flat=True):
            if topics_list:
                all_topics.update(topics_list)
        return Response(sorted(all_topics))

    @action(detail=True, methods=['get'], url_path='transpose')
    def transpose(self, request, pk=None):
        """
        GET /api/v1/songs/{id}/transpose/?semitones=2
        Restituisce il testo con gli accordi trasposti di N semitoni.
        Non modifica il database — è una vista di sola lettura.
        Parametro `semitones`: intero positivo (su) o negativo (giù), default 0.
        """
        song = self.get_object()
        try:
            semitones = int(request.query_params.get('semitones', 0))
        except ValueError:
            return Response({'detail': '`semitones` deve essere un intero.'}, status=status.HTTP_400_BAD_REQUEST)

        lyric_lines = song.lyric_lines.prefetch_related('chords').order_by('order')
        return Response({
            'song_id': str(song.id),
            'title': song.title,
            'original_key': song.key,
            'semitones': semitones,
            'lyric_lines': transpose_song_chords(lyric_lines, semitones),
        })


class LyricLineViewSet(viewsets.ModelViewSet):
    """
    CRUD per le righe testo di una canzone specifica.
    URL annidata: /api/v1/songs/{song_pk}/lines/
    `song_pk` viene dalla URL e garantisce che la riga appartenga alla canzone giusta.
    """
    serializer_class = LyricLineSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtra le righe per canzone (song_pk dalla URL) e per owner."""
        return LyricLine.objects.filter(
            song__owner=self.request.user,
            song_id=self.kwargs['song_pk'],
        ).prefetch_related('chords')

    def perform_create(self, serializer):
        """Collega la nuova riga alla canzone specificata nella URL."""
        song = Song.objects.get(pk=self.kwargs['song_pk'], owner=self.request.user)
        serializer.save(song=song)


class ChordAnnotationViewSet(viewsets.ModelViewSet):
    """
    CRUD per gli accordi di una riga testo specifica.
    URL annidata: /api/v1/songs/{song_pk}/lines/{lyricline_pk}/chords/
    """
    serializer_class = ChordAnnotationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtra gli accordi per riga testo (lyricline_pk dalla URL) e verifica owner."""
        return ChordAnnotation.objects.filter(
            lyric_line__song__owner=self.request.user,
            lyric_line_id=self.kwargs['lyricline_pk'],
        )

    def perform_create(self, serializer):
        """Collega il nuovo accordo alla riga testo specificata nella URL."""
        lyric_line = LyricLine.objects.get(
            pk=self.kwargs['lyricline_pk'],
            song__owner=self.request.user,
        )
        serializer.save(lyric_line=lyric_line)
